# -*- coding: utf-8 -*-
"""
Created on Wed Sep 23 15:02:14 2020

@author: rachana
"""
from tkinter import *
# to perform all operations of excel
from openpyxl import *

#
from tkinter import messagebox
# displays error, warning, info msgs

from tkinter import ttk

from tkcalendar import DateEntry

root = Tk()
root.title("Enter Details")
root.geometry('380x350')


#give path of excel
# unicode error - replace single \ with double \\
wb = load_workbook('C:\\Users\\rachana\\Desktop\\demo.xlsx')
#to make xl sheet in active to all permission
sheet = wb.active 

def excel():
    
    sheet.column_dimensions['A'].width = 20  # width of column = 50
    sheet.cell(row=1, column=1).value = "Name"
    
    sheet.column_dimensions['B'].width = 20
    sheet.cell(row = 1, column = 2).value = "FirstName"
    
    sheet.column_dimensions['C'].width = 20
    sheet.cell(row = 1, column = 3).value = "LastName"
    # insert title to the sheet
    
    sheet.column_dimensions['D'].width = 10
    sheet.cell(row = 1, column = 4).value = "Gender"
    
    sheet.column_dimensions['E'].width = 5
    sheet.cell(row = 1, column = 5).value = "Age"

    sheet.column_dimensions['F'].width = 15
    sheet.cell(row = 1, column = 6).value = "Type"
    
    sheet.column_dimensions['G'].width = 10
    sheet.cell(row = 1, column = 7).value = "Month"
    
    sheet.column_dimensions['H'].width = 10
    sheet.cell(row = 1, column = 8).value = "StartDate"

def insert():
    
    #check if entry field empty
    if (name_field.get() == "" or
        f_field.get() == "" or
        l_field.get() == "" or
        var.get() == "" or
        sp.get() == "" or
        var1.get() == "" or
        month.get() == "" or
        cal.get() == "" ):
    #error msg box    
        print("Empty input")
    
    
    
    current_row = sheet.max_row #max_row gives the no. rows filled 
    current_column = sheet.max_column #
    
    #cell to enter value into cell
    #takes row and column...col is fixed
    #.get() - gets value from 
    sheet.cell(row = current_row + 1, column = 1).value = name_field.get()
    sheet.cell(row = current_row + 1, column = 2).value = f_field.get()
    sheet.cell(row = current_row + 1, column = 3).value = l_field.get()
    sheet.cell(row = current_row + 1, column = 4).value = var.get()
    
    sheet.cell(row = current_row + 1, column = 5).value = sp.get()
    
    sheet.cell(row = current_row + 1, column = 6).value = var1.get()
    
    sheet.cell(row = current_row + 1, column = 7).value = month.get()
    sheet.cell(row = current_row + 1, column = 8).value = cal.get()
    
    wb.save('C:\\Users\\rachana\\Desktop\\demo.xlsx')
    
    
    # messagebox.showerror("Error", "Error message")
    # messagebox.showwarning("Warning", "Warning message")
    messagebox.showinfo("Information", "data inserted successfully")

#func to set focus(cursor)
def focus1(event):
    #set focus on the f_field box
    f_field.focus_set()
 
def focus2(event):
    l_field.focus_set()
    
def focus3(event):
    insert()
  
    
def searchName():
    import page2
    
  
    
excel()



name = Label(root, text='Name').place(x=10,y=10)
fname = Label(root, text='F_Name').place(x=10, y=40)
lname = Label(root, text='L_Name').place(x=10, y=80)



# textfield
name_field = Entry(root)
name_field.place(x=75, y=10)

f_field = Entry(root)
f_field.place(x=75, y=40)

l_field = Entry(root)
l_field.place(x=75, y=80)



#to automatically 
name_field.bind("<Return>", focus1)
f_field.bind("<Return>", focus2)
l_field.bind("<Return>", focus3)



gender = Label(root, text='Gender').place(x=10,y=120)
var = StringVar()
var.set(FALSE)
#variable contains value
r1 = Radiobutton(root, text="Male", variable=var, value="Male")
r1.place(x=70, y=120)

r2 = Radiobutton(root, text="Female", variable=var, value="Female")
r2.place(x=120, y=120)



age = Label(root, text='Age').place(x=10,y=160)
sp = Spinbox(root, from_ =1, to= 10, width=10)
sp.place(x = 75, y=160)



types = Label(root, text='Internship').place(x=10,y=190)
var1 = StringVar()
var1.set(FALSE)

Chkb = Checkbutton(root, text="python", var=var1, onvalue="python")
Chkb.place(x = 70, y = 190)

Chkb1 = Checkbutton(root, text="java", var=var1, onvalue="java")
Chkb1.place(x = 140, y = 190)

Chkb2 = Checkbutton(root, text="HTML", var=var1, onvalue="HTML")
Chkb2.place(x = 190, y = 190)




mont = Label(root, text='Month').place(x=10,y=240)

month = StringVar()

monthChoosen = ttk.Combobox(root, width =10, textvariable = month)
monthChoosen['values'] = ('january', 'february', 'march', 'april', 'may',
                          'june', 'july', 'august', 'september',
                          'october', 'november', 'december')
monthChoosen.place(x= 75, y = 240)
monthChoosen.current(0)



stDate = Label(root, text='Start-Date').place(x=10,y=270)
cal = DateEntry(root, width=12, year=2020, month=9, day=21)
cal.place(x=75, y=270)



submit = Button(root, text="Submit", command=insert)
submit.place(x=120,y=300)

search = Button(root, text="Search", command=searchName)
search.place(x=180,y=300)


root.mainloop()